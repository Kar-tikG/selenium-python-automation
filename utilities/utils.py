import csv
import inspect
import softest
import logging
from openpyxl import workbook, load_workbook

class Utils(softest.TestCase):
    def assert_ele_from_list(self, list_ele, value):
        print(len(list_ele))
        for ele in list_ele:
            print(ele.text)
            #assert ele.text == value
            self.soft_assert(self.assertEqual, ele.text, value)
            if ele.text==value:
                print("test passed")
            else:
                print("test failed")
        self.assert_all()

    def custom_logger(logLevel=logging.DEBUG):
        #Set class/method name from where its called
        logger_name=inspect.stack()[1][3]

        logger=logging.getLogger(logger_name)
        logger.setLevel(logLevel)

        fh=logging.FileHandler("automation.log",mode='w')

        formatter=logging.Formatter("%(asctime)s - %(levelname)s : %(message)s', datefmt='%a, %d %b %Y %H:%M:%SS")

        fh.setFormatter(formatter)

        logger.addHandler(fh)

        return logger

    def read_data_from_excel(file_name, sheet):
        datalist=[]
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column
        for i in range(2, row_ct + 1):
            row=[]
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist

    def read_date_from_csv(file_name):
        datalist=[]         #create an empty list
        csvdata=open(file_name,'r')     #open csv file
        reader=csv.reader(csvdata)      #create csv reader
        next(reader)                    #skip header
        for rows in reader:             #add csv rows to list
            datalist.append(rows)
        return datalist
